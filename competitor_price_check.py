#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""필러 아울렛 경쟁사 가격 조사 (v8 - 진짜 브라우저 지원)
- 안 막히는 사이트: requests (빠름)
- 봇 차단 사이트(meamoshop/acecosm/mjs): Playwright 실제 브라우저로 접속
- 미국 VPN 켠 상태에서 실행하세요.
필요: pip install requests beautifulsoup4 lxml openpyxl cloudscraper curl_cffi
      (playwright/chromium은 이제 선택사항 - meamoshop을 curl_cffi로 처리하므로 보통 불필요)
"""
import csv, html, io, json, os, re, sys, time, datetime, difflib, urllib.parse
import concurrent.futures as cf
try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("필요 패키지 설치:  pip install requests beautifulsoup4 lxml openpyxl cloudscraper playwright")
    print("에러:", e); sys.exit(1)

VERSION = "v9.72"
HERE = os.path.dirname(os.path.abspath(__file__))
PRODUCTS_CSV = os.path.join(HERE, "products.csv")
OUT_DIR = HERE
HISTORY_DIR = os.path.join(HERE, "history")

PRODUCT_SOURCE = "site"
OUR_SITE = "https://filler-outlet.com"
OUR_SITEMAPS = ["https://filler-outlet.com/wp-sitemap.xml"]
OUR_COLLECTIONS = ["https://filler-outlet.com/collections/all-products/"]

USE_BROWSER = False    # v9.45: 브라우저 완전 미사용(requests+curl_cffi로 충분). 창/CAPTCHA 방지, Playwright 불필요.
HEADLESS = False      # 창이 실제로 떠서 스크롤함(meamoshop 무한스크롤 렌더링용). 창 건드리지 마세요.
BROWSER_WAIT_MS = 2500
BROWSER_CAT_DELAY = 4.0  # 브라우저 카테고리 사이 대기(초) - Cloudflare 속도제한 회피 # 페이지 열고 대기(Cloudflare 통과 대기)

MATCH_THRESHOLD = 0.66
MAX_CAT_PAGES = 8  # 브라우저 카테고리당 최대 페이지네이션 페이지 수
REQUEST_TIMEOUT = 25
SLEEP_BETWEEN = 0.4
MAX_WORKERS = 4
PARALLEL_SITES = True   # 7개 경쟁사 카탈로그를 동시에 조사(사이트별 부담은 그대로). 특정 사이트가 삐끗하면 False로.
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,ko;q=0.8",
    "Upgrade-Insecure-Requests": "1",
}
COMPETITORS = [
    {"key": "meamoshop", "base": "https://meamoshop.com", "browser": True, "store_api": True, "wait_selector": ".wd-product, li.product, .product-grid-item", "cat_delay": 4.0,
     "cat_pages": ["https://meamoshop.com/botulinums/","https://meamoshop.com/dermal-fillers/",
                   "https://meamoshop.com/fat-dissolvers/","https://meamoshop.com/skin-boosters/",
                   "https://meamoshop.com/body-fillers/","https://meamoshop.com/lifting-threads/",
                   "https://meamoshop.com/health-boosters/","https://meamoshop.com/numbing-creams/",
                   "https://meamoshop.com/skin-care/","https://meamoshop.com/collagen-stimulators/",
                   "https://meamoshop.com/microneedling/"]},
    {"key": "acecosm", "base": "https://www.acecosm.com", "browser": True,
     "cat_pages": ["https://www.acecosm.com/categories/fillers","https://www.acecosm.com/categories/botulinum-toxins",
                   "https://www.acecosm.com/categories/skin-booster","https://www.acecosm.com/categories/fat-dissolvers",
                   "https://www.acecosm.com/categories/threads","https://www.acecosm.com/categories/body-filler",
                   "https://www.acecosm.com/categories/numbing-creams","https://www.acecosm.com/categories/health-booster",
                   "https://www.acecosm.com/categories/health-supplements","https://www.acecosm.com/categories/skin-care",
                   "https://www.acecosm.com/categories/devices","https://www.acecosm.com/categories/hair-treatment"]},
    {"key": "estaderma", "base": "https://estaderma.com",
     "sitemaps": ["https://estaderma.com/sitemap.xml"],
     "categories": ["https://estaderma.com/shop/fillers/", "https://estaderma.com/shop/toxins/",
                    "https://estaderma.com/shop/skin-boosters/", "https://estaderma.com/shop/fat-dissolving/",
                    "https://estaderma.com/shop/lifting-threads/", "https://estaderma.com/shop/injection/",
                    "https://estaderma.com/shop/numbing-cream/", "https://estaderma.com/shop/disposable/",
                    "https://estaderma.com/shop/devices/", "https://estaderma.com/shop/skincare/",
                    "https://estaderma.com/shop/hair-treatment/", "https://estaderma.com/shop/exosomes/",
                    "https://estaderma.com/shop/curenex/", "https://estaderma.com/shop/stock-clearance/"]},
    {"key": "koreafillerexperts", "base": "https://koreafillerexperts.com", "store_api": True,
     "sitemaps": ["https://koreafillerexperts.com/product-sitemap.xml",
                  "https://koreafillerexperts.com/sitemap_index.xml"]},
    {"key": "derma-solution", "base": "https://derma-solution.com", "store_api": True,
     "sitemaps": ["https://derma-solution.com/product-sitemap.xml"]},
    {"key": "mjsmedicals", "base": "https://www.mjsmedicals.com", "browser": True, "wait_selector": "li.product, .wd-product, .product-grid-item",
     "cat_pages": ["https://www.mjsmedicals.com/collections/dermal-filler/",
                   "https://www.mjsmedicals.com/collections/botulinum-toxin/",
                   "https://www.mjsmedicals.com/collections/skinbooster/",
                   "https://www.mjsmedicals.com/collections/lifting-thread/",
                   "https://www.mjsmedicals.com/collections/fat-dissolver/",
                   "https://www.mjsmedicals.com/collections/exosome/",
                   "https://www.mjsmedicals.com/shop/"]},
    {"key": "fillerhouse", "base": "https://fillerhouse.com",
     "sitemaps": ["https://fillerhouse.com/sitemap.xml"],
     "categories": ["https://fillerhouse.com/shop/"]},
]
COMP_KEYS = [c["key"] for c in COMPETITORS]
SITE_STATUS = {}
URL_FAIL = {}   # (comp_key, our_norm) -> True : 수동보정 URL이 있었는데 실패해서 자동값 폴백/빈칸된 것
BLOCK_MARKERS = ("attention required! | cloudflare", "just a moment...", "access-denied",
                 "site is down for maintenance", "checking your browser before",
                 "verifying you are human", "ddos-guard")

_session = requests.Session(); _session.headers.update(HEADERS)
_scraper = None
def _get_scraper():
    global _scraper
    if _scraper is None:
        try:
            import cloudscraper
            _scraper = cloudscraper.create_scraper()
        except Exception:
            _scraper = False
    return _scraper

_ccreq = None
def _get_ccreq():
    """curl_cffi(크롬 TLS 지문 흉내)를 지연 로드. 없으면 False."""
    global _ccreq
    if _ccreq is None:
        try:
            from curl_cffi import requests as ccr
            _ccreq = ccr
        except Exception:
            _ccreq = False
    return _ccreq

def cfetch(url, tries=1):
    """curl_cffi로 가져오기(크롬 TLS 흉내). Cloudflare가 순수 requests를 막는 사이트(meamoshop 등)를 브라우저 없이 뚫는다. 실패시 None."""
    cc = _get_ccreq()
    if not cc:
        return None
    for imp in ("chrome", "chrome120", "chrome123", "safari"):
        for _k in range(tries):
            try:
                r = cc.get(url, impersonate=imp, timeout=REQUEST_TIMEOUT)
                if r.status_code == 200 and r.text:
                    return r.text
            except Exception:
                pass
            if _k + 1 < tries:
                time.sleep(1.0)
    return None

def fetch(url, tries=3):
    for i in range(tries):
        try:
            r = _session.get(url, timeout=REQUEST_TIMEOUT)
            if r.status_code == 429:
                time.sleep(3 * (i + 1)); continue
            if r.status_code == 200:
                return r.text
            if r.status_code in (403, 503):
                sc = _get_scraper()
                if sc:
                    try:
                        r2 = sc.get(url, timeout=REQUEST_TIMEOUT)
                        if r2.status_code == 200:
                            return r2.text
                    except Exception:
                        pass
                return None
            return None
        except requests.RequestException:
            time.sleep(1.5 * (i + 1))
    # 순수 requests가 실패/차단 → curl_cffi(크롬 TLS)로 마지막 시도(브라우저 없이 Cloudflare 통과)
    ct = cfetch(url)
    if ct is not None and not is_blocked(ct):
        return ct
    return None

# ---- 진짜 브라우저 (Playwright) ----
_pw = None; _browser = None; _bpage = None
def _browser_page():
    global _pw, _browser, _bpage
    if _bpage is None:
        from playwright.sync_api import sync_playwright
        _pw = sync_playwright().start()
        args = ["--disable-blink-features=AutomationControlled", "--no-sandbox",
                "--disable-features=IsolateOrigins,site-per-process"]
        # 프로필 저장형 브라우저: 한 번 Cloudflare(사람 확인)를 통과하면 그 쿠키(cf_clearance)를
        # 프로필 폴더에 저장 → 다음 실행 때 재사용해서 재검사 빈도를 크게 줄임.
        profile_dir = os.path.join(HERE, ".browser_profile")
        try:
            os.makedirs(profile_dir, exist_ok=True)
        except Exception:
            pass
        ctx = _pw.chromium.launch_persistent_context(
            profile_dir, headless=HEADLESS, args=args,
            user_agent=HEADERS["User-Agent"], locale="en-US",
            viewport={"width": 1366, "height": 900})
        ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
                            "window.chrome = window.chrome || {runtime: {}};")
        _browser = ctx  # persistent context가 브라우저 역할(닫을 땐 ctx.close())
        _bpage = ctx.pages[0] if ctx.pages else ctx.new_page()
    return _bpage

def browser_fetch(url, tries=2, wait_selector=None):
    html = None
    for i in range(tries):
        try:
            page = _browser_page()
            page.goto(url, timeout=40000, wait_until="domcontentloaded")
            page.wait_for_timeout(2000)
            # Cloudflare 등 봇검사면 통과될 때까지 대기(창모드: 사람이 체크박스 클릭 / 또는 자동 통과)
            challenged = False
            for _ in range(30):                 # 최대 ~60초
                html = page.content()
                if not is_blocked(html):
                    break
                challenged = True
                page.wait_for_timeout(2000)
            if is_blocked(html):
                continue                          # 못 뚫음 → 재시도
            if wait_selector:
                try:
                    page.wait_for_selector(wait_selector, timeout=10000)
                except Exception:
                    pass
            try:
                page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            # 무한스크롤 + "더보기(Load more)" 버튼 둘 다 처리: 상품이 안 늘 때까지 반복
            try:
                load_more_js = ("() => { const b = document.querySelector("
                                "'a.wd-load-more:not(.wd-hide), .wd-load-more:not(.wd-hide), "
                                "a.load-more, .load-more-button a, button.load-more, .wd-products-load-more'); "
                                "if (b && b.offsetParent !== null) { b.scrollIntoView({block:'center'}); "
                                "b.click(); return true; } return false; }")
                count_js = ("document.querySelectorAll('.wd-product, li.product, "
                            ".product-grid-item, .product').length")
                prev_count, stable = -1, 0
                for _ in range(40):
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    page.wait_for_timeout(800)
                    try:
                        clicked = page.evaluate(load_more_js)
                    except Exception:
                        clicked = False
                    page.wait_for_timeout(1600 if clicked else 400)
                    count = page.evaluate(count_js)
                    if count == prev_count and not clicked:
                        stable += 1
                        if stable >= 2:
                            break
                    else:
                        stable = 0
                    prev_count = count
                page.evaluate("window.scrollTo(0, 0)")
                page.wait_for_timeout(600)
            except Exception:
                pass
            return page.content()
        except Exception:
            time.sleep(2)
    return html

def close_browser():
    global _pw, _browser, _bpage
    try:
        if _browser: _browser.close()   # persistent context 저장 후 종료
    except Exception:
        pass
    try:
        if _pw: _pw.stop()
    except Exception:
        pass
    _pw = _browser = _bpage = None

def _sync_browser_cookies():
    """브라우저(Playwright)가 Cloudflare를 통과하며 받은 쿼키(cf_clearance 등)를
    requests 세션(_session)에 복사한다. 이후 수동보정 상세페이지를
    브라우저 순차폴백(느림) 대신 requests 병렬(빠름)로 받을 수 있음.
    (meamoshop이 Store API를 무접속 조회하던 것과 같은 원리). 반환: 복사한 쿼키 개수."""
    global _browser
    if _browser is None:
        return 0
    try:
        cookies = _browser.cookies()      # persistent context의 전체 쿼키
    except Exception:
        return 0
    n = 0
    for c in cookies or []:
        name = c.get("name"); val = c.get("value")
        if not name:
            continue
        try:
            _session.cookies.set(name, val, domain=c.get("domain"), path=c.get("path") or "/")
            n += 1
        except Exception:
            pass
    return n

# 브라우저(진짜 크롬) 안에서 여러 상세페이지를 '동시에' fetch하여 가격/재고를 뽑는 JS.
# 브라우저는 Cloudflare를 통과했고 TLS지문도 크롬이라, requests가 못 받던 mjs HTML을 정상적으로 받는다.
_BATCH_PRICE_JS = r"""async (urls) => {
  const out = {};
  const CONC = 6;
  let idx = 0;
  function priceFromDoc(doc){
    doc.querySelectorAll('.related, .related-products, .up-sells, .upsells, .cross-sells, .cross-sell, .products-carousel, .wd-slider, .single-related, [class*="tooltip"], .product-navigation').forEach(e=>e.remove());
    const wd = doc.querySelector('.wd-single-price');
    if (wd){
      const ins = wd.querySelector('ins .amount, ins .woocommerce-Price-amount, ins');
      const el = ins || wd;
      const m = (el.textContent||'').match(/\$\s*([\d][\d,]*(?:\.\d+)?)/);
      if (m){ const v=parseFloat(m[1].replace(/,/g,'')); if(v>0) return v; }
    }
    const vf = doc.querySelector('form.variations_form, [data-product_variations]');
    if (vf){
      const raw = vf.getAttribute('data-product_variations');
      if (raw && raw!=='false' && raw!=='[]'){
        try{ const arr=JSON.parse(raw); const ps=[];
          for (const v of arr){ let dp=(v&&(v.display_price!=null?v.display_price:v.price)); if(dp!=null&&dp!==''){ const f=parseFloat(dp); if(f>0) ps.push(f); } }
          if (ps.length) return Math.min.apply(null, ps);
        }catch(e){}
      }
    }
    for (const sc of doc.querySelectorAll('script[type="application/ld+json"]')){
      try{ let data=JSON.parse(sc.textContent||''); let stack=Array.isArray(data)?data.slice():[data];
        for (const n of stack.slice()){ if(n&&Array.isArray(n['@graph'])) stack=stack.concat(n['@graph']); }
        for (const n of stack){ if(!n||typeof n!=='object') continue; let t=n['@type']; t=Array.isArray(t)?(t[0]||''):(t||'');
          if(String(t).toLowerCase()==='product'){ let off=n.offers; if(Array.isArray(off)) off=off[0]||{};
            if(off){ let pr=off.price||off.lowPrice; if(pr){ const f=parseFloat(String(pr).replace(/,/g,'')); if(f>0) return f; } } } }
      }catch(e){}
    }
    const meta = doc.querySelector('meta[property="product:price:amount"], meta[property="og:price:amount"], meta[itemprop="price"]');
    if (meta && meta.getAttribute('content')){ const f=parseFloat(meta.getAttribute('content').replace(/,/g,'')); if(f>0) return f; }
    return null;
  }
  function stockFromDoc(doc){
    const av = doc.querySelector('link[itemprop="availability"], meta[itemprop="availability"], meta[property="product:availability"], meta[property="og:availability"]');
    if (av){ const v=((av.getAttribute('href')||'')+' '+(av.getAttribute('content')||'')).toLowerCase(); if(v.includes('outofstock')||v.includes('out_of_stock')||v.includes('sold')) return false; }
    if (doc.querySelector('p.stock.out-of-stock, .stock.out-of-stock')) return false;
    return true;
  }
  async function worker(){
    while (idx < urls.length){
      const u = urls[idx++];
      try{
        const r = await fetch(u, {headers:{'Accept':'text/html'}, credentials:'include'});
        const t = await r.text();
        const doc = new DOMParser().parseFromString(t, 'text/html');
        out[u] = {price: priceFromDoc(doc), instock: stockFromDoc(doc)};
      }catch(e){ out[u] = {price:null, instock:true, err:String(e)}; }
    }
  }
  const ws=[]; for(let k=0;k<CONC;k++) ws.push(worker());
  await Promise.all(ws);
  return out;
}"""

def _browser_batch_prices(comp_key, tasks, chunk=40, time_budget=150):
    """브라우저 안에서 상세페이지들을 병렬 fetch → {url: {'price':float|None,'instock':bool}}.
    requests가 실패한 mjs 등을 순차 브라우저(24분) 대신 동시처리(수십초)로 해결."""
    comp = COMP_BY_KEY.get(comp_key, {})
    base = comp.get("base")
    urls = [u for (_on, u) in tasks]
    if not urls or not base:
        return {}
    try:
        page = _browser_page()
        page.goto(base + "/", timeout=40000, wait_until="domcontentloaded")
        page.wait_for_timeout(1200)
        for _ in range(30):                 # Cloudflare 통과 대기
            if not is_blocked(page.content()):
                break
            page.wait_for_timeout(2000)
        if is_blocked(page.content()):
            return {}
    except Exception:
        return {}
    result, t0 = {}, time.time()
    for i in range(0, len(urls), chunk):
        if time.time() - t0 > time_budget:
            break
        part = urls[i:i + chunk]
        try:
            data = page.evaluate(_BATCH_PRICE_JS, part)
        except Exception:
            data = None
        if isinstance(data, dict):
            result.update(data)
        page.wait_for_timeout(150)
    return result

# 브라우저 안에서 Store API를 slug(주소조각)로 조회 → 특정 제품만 JSON으로 빠르게 받기.
# (slug 지도에 없던 meamoshop 수동보정 제품을, 페이지 통째 렌더 없이 초 단위로 확보하는 용도)
_STORE_API_SLUG_JS = r"""async (slugs) => {
  const out = {}; const CONC = 6; let i = 0;
  async function worker(){
    while (i < slugs.length){
      const s = slugs[i++];
      try{
        const r = await fetch(location.origin + '/wp-json/wc/store/v1/products?slug=' + encodeURIComponent(s),
                              {headers:{'Accept':'application/json'}, credentials:'include'});
        const j = await r.json();
        if (Array.isArray(j) && j.length){
          const p = j[0]; const pp = p.prices || {};
          let price = null;
          if (pp.price != null && pp.price !== '') price = parseFloat(pp.price) / Math.pow(10, parseInt(pp.currency_minor_unit || 2));
          out[s] = {price: (price && price > 0 ? price : null), instock: (p.is_in_stock !== false)};
        } else { out[s] = {price: null}; }
      }catch(e){ out[s] = {price: null, err: String(e)}; }
    }
  }
  const ws = []; for (let k = 0; k < CONC; k++) ws.push(worker());
  await Promise.all(ws);
  return out;
}"""

def _browser_store_api_by_slug(comp_key, slugs, chunk=40):
    """브라우저(Cloudflare 통과 상태)에서 Store API를 slug로 조회 → {slug: {'price':float|None,'instock':bool}}.
    개별 페이지 렌더(느림) 대신 이걸 쓰면 버림 없이 전부 빠르게 확보된다."""
    comp = COMP_BY_KEY.get(comp_key, {}); base = comp.get("base")
    slugs = [s for s in slugs if s]
    if not slugs or not base:
        return {}
    try:
        page = _browser_page()
        page.goto(base + "/", timeout=40000, wait_until="domcontentloaded")   # 원점(origin)+CF 맞추기
        page.wait_for_timeout(1000)
        for _ in range(30):
            if not is_blocked(page.content()):
                break
            page.wait_for_timeout(2000)
        if is_blocked(page.content()):
            return {}
    except Exception:
        return {}
    result = {}
    for i in range(0, len(slugs), chunk):
        part = slugs[i:i + chunk]
        try:
            data = page.evaluate(_STORE_API_SLUG_JS, part)
        except Exception:
            data = None
        if isinstance(data, dict):
            result.update(data)
    return result

def is_blocked(html):
    if not html:
        return False
    low = html[:5000].lower()
    return any(m in low for m in BLOCK_MARKERS)

_DASH_RE = re.compile("[\u2010\u2011\u2012\u2013\u2014\u2015\u2212\uFE58\uFE63\uFF0D]")

def _fix_dash(s):
    """en-dash(–)·em-dash(—)·minus(−) 등 모든 대시 변형을 일반 하이픈(-)으로 통일."""
    return _DASH_RE.sub("-", s) if s else s

def normalize(name):
    s = _fix_dash(name.lower().replace("＋", "+"))
    s = re.sub(r"(\d),(\d)", r"\1\2", s)
    s = re.sub(r"\b(with )?lidocaine\b", "", s)
    s = re.sub(r"\b(inj|injection)\b", "", s)
    s = re.sub(r"[\(\)\[\]{}.,/#]", " ", s)
    s = s.replace("units", "u").replace("unit", "u").replace("ui", "u")
    return re.sub(r"\s+", " ", s).strip()

def unit_sig(s):
    return {m.group(1) + m.group(2) for m in re.finditer(r"(\d+)\s*(u|ml|cc|mg|g|a)\b", s)}

def units_ok(a, b):
    return not (a and b and not (a & b))

def plus_ok(a, b):
    """'Plus'(리도카인 버전) 유무가 다르면 서로 다른 제품 → 매칭 금지.
    예: 'elasty d'(리도카인X) ↔ 'elasty d plus'(리도카인O) 는 매칭 안 됨."""
    return ("plus" in a.split()) == ("plus" in b.split())

# 브랜드가 아닌 '흔한 수식어/제형' 단어들 — 이 단어만 같은 건 매칭 근거로 인정 안 함
_GENERIC = {
    "volume","balance","intense","soft","deep","fine","ultra","light","lite","forte",
    "plus","hard","strong","mild","premium","gold","silver","pro","max","mini","sub","subq",
    "lido","lidocaine","with","without","new","original","classic","fresh",
    "cream","gel","mask","pack","booster","boosters","solution","serum","ampoule","ampoules",
    "filler","fillers","toxin","toxins","thread","threads","set","kit","bundle","combo","program",
    "injection","body","face","facial","skin","skincare","daily","eye","eyes","lip","lips",
    "cleanser","toner","essence","recovery","repair","rejuvenating","hydro","exo","tone","up",
    "care","line","the","and","for","box","pcs","vial","vials","syringe","syringes","booster",
}

def _is_brand_token(t):
    """흔한 수식어가 아니고 4글자 이상인 '브랜드성' 단어인지(단일 공통단어 강매칭 판정용).
    숫자로 시작하는 토큰(100u·50iu·10ml 등 용량/단위)은 브랜드가 아니므로 제외."""
    core = re.sub(r"[^a-z0-9]", "", t)
    return (len(core) >= 4 and core not in _GENERIC
            and not core.isdigit() and not core[:1].isdigit())

def _distinctive_ok(a, b):
    """우리 제품의 '브랜드성 토큰'(흔한 수식어 제외)이 상대 이름에 하나도 없으면 매칭 금지.
    예: 'belotero volume' 의 브랜드 토큰은 'belotero' → 상대에 belotero 없으면(‘metoo volume’ 등) 매칭 안 함.
    브랜드 토큰이 아예 없는 제품은 판단 불가라 통과."""
    ta = set()
    for t in a.split():
        core = re.sub(r"[^a-z0-9]", "", t)
        if len(core) >= 3 and core not in _GENERIC and not core.isdigit() and not core[:1].isdigit():
            ta.add(core)   # 숫자로 시작하는 용량토큰(100u 등)은 브랜드 아님 → 제외
    if not ta:
        return True
    bc = re.sub(r"[^a-z0-9]", "", b)
    return any(t in bc for t in ta)

def pack_count(s):
    """제품명에서 묶음 수량 추출: 1x1ml->1, 2x1ml->2, 2ea/2 syringes->2. 없으면 1."""
    m = re.search(r"(\d+)\s*x\s*\d", s)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(\d+)\s*(ea|pcs|pieces|syringes?|vials?|packs?|tubes?)\b", s)
    if m:
        return int(m.group(1))
    return 1

def pack_count_explicit(s):
    """이름에 수량이 '명시'돼 있으면 그 수, 없으면 None(1x1ml/2 syringes 등)."""
    m = re.search(r"(\d+)\s*x\s*\d", s)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(\d+)\s*(ea|pcs|pieces|syringes?|vials?|packs?|tubes?)\b", s)
    if m:
        return int(m.group(1))
    return None

def pack_ok(a, b):
    """양쪽 다 수량이 명시된 경우에만 수량 일치를 요구(1실린지 vs 2실린지 혼동 방지).
    한쪽이라도 수량 표기가 없으면 막지 않음 → 박스/팩 등 기존 매칭은 그대로 유지."""
    ca, cb = pack_count_explicit(a), pack_count_explicit(b)
    if ca is None or cb is None:
        return True
    return ca == cb

def match_score(a, b):
    """이름 매칭 점수. 짧은 쪽 핵심 단어가 긴 쪽에 다 들어있으면(예: 'puri lips' ⊆ 'aeterderm puri lips') 강매칭."""
    if not a or not b:
        return 0.0
    ta = {t for t in a.split() if len(t) > 1}
    tb = {t for t in b.split() if len(t) > 1}
    ratio = difflib.SequenceMatcher(None, a, b).ratio()
    # 브랜드 표기 차이 흡수: 공백 제거 후 짧은 쪽이 긴 쪽의 앞부분과 일치하면 강매칭
    # (예: 'lipo lab ppc' vs 'lipolab ppc solution' → 'lipolabppc' 가 'lipolabppcsolution' 의 접두)
    ca = re.sub(r"[^a-z0-9]", "", a); cb = re.sub(r"[^a-z0-9]", "", b)
    if ca and cb:
        _short, _long = (ca, cb) if len(ca) <= len(cb) else (cb, ca)
        if len(_short) >= 6 and _long.startswith(_short):
            return max(ratio, 0.9)
    if ta and tb:
        inter = ta & tb
        smaller = min(len(ta), len(tb))
        if len(inter) == smaller and abs(len(ta) - len(tb)) <= 1 \
           and (len(inter) >= 2 or (len(inter) == 1 and _is_brand_token(next(iter(inter))))):
            return max(ratio, 0.9)
    return ratio

def _amount_from(el):
    """요소 텍스트에서 첫 유효 금액(>0) 추출. "$ 35.00"(기호·숫자 분리)·"35.00"·"1,234.50" 모두 처리."""
    if el is None:
        return None
    txt = el.get_text(" ", strip=True)
    for m in re.finditer(r"([\d][\d,]*(?:\.\d+)?)", txt):
        try:
            v = float(m.group(1).replace(",", ""))
        except ValueError:
            continue
        if v > 0:
            return v
    return None

def parse_price_from_html(html):
    soup = BeautifulSoup(html, "lxml")
    # [최우선] WoodMart+Elementor(mjsmedicals 등) 본품가는 .wd-single-price 위젯 안에 있다.
    #  통화기호($)와 숫자(35.00)가 다른 태그로 분리돼 "$ 35.00"으로 오고, 페이지엔 $0.00 placeholder가
    #  여럿 앞서 있어서 예전엔 못 집었음(Elasty 262/None 버그). junk 제거 前 원본 soup에서 먼저 뽑는다.
    _wsp = soup.select_one(".wd-single-price")
    if _wsp is not None:
        _ins = _wsp.select_one("ins .woocommerce-Price-amount, ins .amount, ins")
        _amt = _wsp.select_one(".woocommerce-Price-amount, .amount, bdi")
        _v = _amount_from(_ins) or _amount_from(_amt) or _amount_from(_wsp)
        if _v:
            return _v
    # 이전/다음 상품 미리보기(툴팁)·추천·연관 상품 그리드의 가격 제거 → 본문 가격만 남김
    for junk in soup.select('.product-navigation, [class*="tooltip"], .related, '
                            '.related-products, .up-sells, .upsells, .cross-sells, '
                            '.cross-sell, ul.products, .wd-products, .single-related, '
                            '.products-carousel, .wd-slider'):
        try:
            junk.decompose()
        except Exception:
            pass
    # 커스텀 사이트 본품 가격 우선 위치(acecosm 등): 관련상품/장바구니 말고 진짜 본품 가격
    # acecosm(커스텀 플랫폼): 한 페이지에 관련상품 가격(.money/.price-box)이 여러 개 섞여
    #  첫 요소가 엉뚱한 값($190)일 수 있음 → 페이지에 박힌 본품 JSON "price":N 을 사용(정확)
    if soup.select_one(".price-box") is not None:
        mm = re.search(r'"price"\s*:\s*"?(\d+(?:\.\d+)?)"?', html)
        if mm and float(mm.group(1)) > 0:
            return float(mm.group(1))
    # WoodMart 테마(mjsmedicals 등): 상세페이지 본품 가격이 .wd-single-price 안에만 정확히 있음.
    #  이 위치를 안 보면 페이지에 섞인 수량할인표·연관상품 가격($121.8/262 등)을 잘못 집음(Elasty D/F/G 262버그).
    #  세일가(ins)가 있으면 그것을, 없으면 첫 가격을 본품가로 사용.
    wd_single = soup.select_one(".wd-single-price")
    if wd_single is not None:
        ins = wd_single.select_one("ins .amount, ins .woocommerce-Price-amount, ins")
        target = ins if ins is not None else wd_single
        mm = re.search(r"\$\s*([\d][\d,]*(?:\.\d+)?)", target.get_text(" ", strip=True))
        if mm:
            v = float(mm.group(1).replace(",", ""))
            if v > 0:
                return v
    # 옵션(변형) 상품: 본품가가 $0.00으로 뜨는 경우가 많음 → 변형 가격(>0) 중 최솟값을 대표가로
    vf = soup.select_one("form.variations_form, [data-product_variations]")
    if vf:
        raw = vf.get("data-product_variations")
        if raw and raw not in ("false", "[]"):
            try:
                vps = []
                for v in json.loads(raw):
                    if isinstance(v, dict):
                        dp = v.get("display_price", v.get("price"))
                        if dp not in (None, ""):
                            fv = float(dp)
                            if fv > 0:
                                vps.append(fv)
                if vps:
                    return min(vps)
            except Exception:
                pass
    for tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(tag.string or "")
        except Exception:
            continue
        stack = data if isinstance(data, list) else [data]
        for node in list(stack):
            if isinstance(node, dict) and isinstance(node.get("@graph"), list):
                stack.extend(node["@graph"])
        for node in stack:
            if not isinstance(node, dict):
                continue
            t = node.get("@type")
            t = t if isinstance(t, str) else (t[0] if isinstance(t, list) and t else "")
            if str(t).lower() == "product":
                offers = node.get("offers")
                if isinstance(offers, list):
                    offers = offers[0] if offers else {}
                if isinstance(offers, dict):
                    pr = offers.get("price") or offers.get("lowPrice")
                    if pr:
                        try:
                            f = float(str(pr).replace(",", ""))
                            if f > 0:
                                return f
                        except ValueError:
                            pass
    for sel in ('meta[property="product:price:amount"]', 'meta[property="og:price:amount"]',
                'meta[itemprop="price"]'):
        el = soup.select_one(sel)
        if el and el.get("content"):
            try:
                f = float(el["content"].replace(",", ""))
                if f > 0:
                    return f
            except ValueError:
                pass
    for el in soup.select('ins .amount, .price ins .amount, ins .woocommerce-Price-amount'):
        mm = re.search(r"\$\s*([\d][\d,]*(?:\.\d+)?)", el.get_text(" ", strip=True))
        if mm and float(mm.group(1).replace(",", "")) > 0:
            return float(mm.group(1).replace(",", ""))
    for el in soup.select('.woocommerce-Price-amount, .price .amount, .price, .product-price, '
                          '[class*="price"], [class*="amount"]'):
        mm = re.search(r"\$\s*([\d][\d,]*(?:\.\d+)?)", el.get_text(" ", strip=True))
        if mm and float(mm.group(1).replace(",", "")) > 0:
            return float(mm.group(1).replace(",", ""))
    return None

_OOS_JSON = re.compile(r'"availability"\s*:\s*"[^"]*(OutOfStock|SoldOut|out_of_stock)', re.I)
def is_out_of_stock(html):
    """상품 페이지가 품절인지(구조적 신호 기반, 오탐 최소화)."""
    if not html:
        return False
    low = html.lower()
    if _OOS_JSON.search(html) or "schema.org/outofstock" in low or "schema.org/soldout" in low:
        return True
    soup = BeautifulSoup(html, "lxml")
    for el in soup.select('link[itemprop="availability"], meta[itemprop="availability"], '
                          'meta[property="product:availability"], meta[property="og:availability"]'):
        val = ((el.get("href") or "") + " " + (el.get("content") or "")).lower()
        if "outofstock" in val or "out_of_stock" in val or "sold" in val:
            return True
    if soup.select_one("p.stock.out-of-stock, .stock.out-of-stock, "
                       ".woocommerce-variation-availability .out-of-stock"):
        return True
    return False

def parse_name_from_html(html):
    soup = BeautifulSoup(html, "lxml")
    # h1 제품명은 그대로 사용 → 'Rejuran Tone-Up Booster'처럼 하이픈 포함 이름을 안 자름
    for sel in ("h1.product_title", "h1.product-title", "h1.entry-title"):
        el = soup.select_one(sel)
        if el:
            txt = el.get_text(" ", strip=True)
            if txt and len(txt) >= 2:
                return txt
    # 폴백: <title>/og:title 은 ' - 사이트명' / ' | 사이트명' 접미사만 제거(양옆 공백 있는 구분자만)
    for sel in ("meta[property='og:title']", "title"):
        el = soup.select_one(sel)
        if el:
            txt = el.get("content") if el.name == "meta" else el.get_text()
            if txt:
                return re.sub(r"\s+[-–|]\s+[^-–|]*$", "", txt).strip()
    return ""

def parse_variations(html):
    """WooCommerce 옵션(변형) 상품의 옵션별 (라벨, 가격) 목록. 없으면 []"""
    soup = BeautifulSoup(html, "lxml")
    el = soup.select_one("form.variations_form, [data-product_variations]")
    if not el:
        return []
    raw = el.get("data-product_variations")
    if not raw or raw in ("false", "[]"):
        return []
    try:
        arr = json.loads(raw)
    except Exception:
        return []
    out = []
    for v in arr:
        if not isinstance(v, dict):
            continue
        attrs = v.get("attributes") or {}
        label = " ".join(str(x) for x in attrs.values() if x).strip()
        price = v.get("display_price", v.get("price"))
        try:
            price = float(price)
        except (TypeError, ValueError):
            continue
        if label and price > 0:
            out.append((label, price))
    return out

def get_sitemap_urls(sitemap_urls, fetcher=fetch):
    found, seen, queue = [], set(), list(sitemap_urls)
    while queue:
        sm = queue.pop(0)
        if sm in seen:
            continue
        seen.add(sm)
        xml = fetcher(sm)
        if not xml or is_blocked(xml):
            continue
        for loc in re.findall(r"<loc>\s*([^<\s]+)\s*</loc>", xml):
            if loc.endswith(".xml") and "sitemap" in loc.lower():
                queue.append(loc)
            else:
                found.append(loc)
    return found

def product_urls_from_pages(pages, base, fetcher=fetch,
                            patterns=("/product", "/products/", "/shop/", "/categories/")):
    urls = []
    for pg in pages:
        html = fetcher(pg); time.sleep(SLEEP_BETWEEN)
        if not html or is_blocked(html):
            continue
        soup = BeautifulSoup(html, "lxml")
        for a in soup.select("a[href]"):
            href = a["href"]
            if any(k in href for k in patterns):
                urls.append(urllib.parse.urljoin(base, href))
    return list(dict.fromkeys(urls))

def slug_of(url):
    path = urllib.parse.urlparse(url).path.rstrip("/")
    return path.split("/")[-1].replace("-", " ")

def parse_catalog(html, include_sets=False):
    """메인 상품 그리드 + 일반 링크(메뉴/추천 등) 둘 다 긁어서 합침 → 렌더된 건 최대한 다 수집."""
    soup = BeautifulSoup(html, "lxml")
    price_sel = ".woocommerce-Price-amount, .money, .amount, .price, [class*='price']"
    skip = ("quick view","add to cart","select options","read more","compare",
            "add to compare","add to wishlist","shop now","view all","home",
            "popularity","average rating","newness","price: low to high",
            "price: high to low","default sorting","all","botulinum toxins",
            "dermal fillers","skin boosters","fat dissolvers","lifting threads")
    def is_set(nm):
        return bool(re.search(r"\b(set|kit|bundle|combo)\b", nm.lower()))
    def price_of(el):
        pe = el.select_one(price_sel)
        if pe:
            m = re.search(r"\$\s*([\d][\d,]*(?:\.\d+)?)", pe.get_text(" ", strip=True))
            if m:
                v = float(m.group(1).replace(",", ""))
                if 0 < v < 100000:
                    return v
        return None
    def _card_oos(card):
        cls = " ".join(card.get("class", []) or []).lower()
        if "outofstock" in cls or "out-of-stock" in cls or "sold-out" in cls or "soldout" in cls:
            return True
        for b in card.select('.out-of-stock, .outofstock, .sold-out, .soldout, [class*="sold"], '
                             '.wd-product-label, .product-label, .stock, .badge'):
            if re.search(r"out of stock|sold out|품절", b.get_text(" ", strip=True), re.I):
                return True
        return False

    def _card_url(card):
        for sel in ("a.wd-entities-title", ".wd-entities-title a", ".woocommerce-loop-product__title a",
                    "h2 a", "h3 a", ".product-title a", "a.product-image-link", "a[href]"):
            el = card.select_one(sel)
            if el and el.get("href"):
                h = el["href"]
                if h and not h.startswith("#") and "javascript" not in h:
                    return h
        return ""

    items = []
    # A) 메인 상품 그리드(카드 단위) — 이름이 정확
    grid, best = None, 0
    for g in soup.select("ul.products, div.products.wd-products, .wd-products, div.products, "
                         ".wc-block-grid__products, .product-grid, .products-grid"):
        n = len(g.select("li.product, .wd-product, .product-grid-item, .product"))
        if n > best:
            best, grid = n, g
    if grid is not None and best >= 3:
        for card in grid.select("li.product, .wd-product, .product-grid-item, .product"):
            nel = card.select_one(".wd-entities-title a, .wd-entities-title, .woocommerce-loop-product__title, "
                                  "h2 a, h3 a, h2, h3, .product-title a, .product-title, a[href]")
            name = nel.get_text(" ", strip=True) if nel else ""
            if not name or len(name) < 3:
                a_ = card.select_one("a[href]")
                if a_ and a_.get("href"):
                    name = slug_of(a_["href"])   # 카드 이름이 비면 URL 슬러그로
            if not name or len(name) < 3 or name.lower() in skip or (is_set(name) and not include_sets):
                continue
            pr = price_of(card)
            if pr is not None:
                items.append((name, pr, _card_oos(card), _card_url(card)))

    # B) 일반 링크+주변 가격 (메뉴/추천/무엇이든 렌더된 것) — 예전 방식
    for a in soup.select("a[href]"):
        name = a.get_text(" ", strip=True); href = a.get("href", "")
        if not name or len(name) < 3 or name.lower() in skip or (is_set(name) and not include_sets):
            continue
        if (not href) or href.startswith("#") or "javascript" in href or href.startswith("mailto") \
           or any(k in href for k in ("/cart","/account","/wishlist","/login","/compare","/register")):
            continue
        box = a; pr = None
        for _ in range(4):
            box = box.parent
            if box is None:
                break
            pr = price_of(box)
            if pr is not None:
                break
        if pr is not None:
            items.append((name, pr, False, href))

    seen, out = set(), []
    for nm, pr, oos, url in items:
        key = normalize(nm)
        if not key or key in seen:
            continue
        seen.add(key); out.append((nm, pr, not oos, url))
    return out

def _next_page_url(html, cur_url, visited):
    """WooCommerce 페이지네이션의 다음 페이지 URL(없으면 None)."""
    soup = BeautifulSoup(html, "lxml")
    for a in soup.select("a.next.page-numbers, a.next, .woocommerce-pagination a.next, "
                         ".pagination a.next, nav.pagination a.next"):
        href = a.get("href")
        if href:
            u = urllib.parse.urljoin(cur_url, href)
            if u not in visited:
                return u
    m = re.search(r"/page/(\d+)/?", cur_url)
    cur_n = int(m.group(1)) if m else 1
    cand = []
    for a in soup.select("a.page-numbers, .page-numbers a, .woocommerce-pagination a, .pagination a"):
        href = a.get("href") or ""
        mm = re.search(r"/page/(\d+)/?", href)
        if mm:
            n = int(mm.group(1))
            if n > cur_n:
                cand.append((n, urllib.parse.urljoin(cur_url, href)))
    cand.sort()
    for _n, u in cand:
        if u not in visited:
            return u
    return None

def _match_catalog(catalog, product_norms, base):
    """catalog: [(norm, unit_sig, price, in_stock, url)] → 우리 제품별 최적 매칭(옵션 중복 방지 포함).
    매칭된 카탈로그 항목의 실제 상세URL을 함께 저장(없으면 base). → 자동매칭도 클릭 검증 가능."""
    matched = {}
    for our_norm in product_norms:
        osig = unit_sig(our_norm)
        best, bidx, bprice, bstock, burl = 0.0, None, None, True, base
        for idx, item in enumerate(catalog):
            nn, ssig, pr, in_stock = item[0], item[1], item[2], item[3]
            curl = item[4] if len(item) > 4 else base
            if not nn or not units_ok(osig, ssig) or not plus_ok(our_norm, nn) \
               or not pack_ok(our_norm, nn) or not _distinctive_ok(our_norm, nn):
                continue
            ratio = match_score(our_norm, nn)
            if ratio > best:
                best, bidx, bprice, bstock, burl = ratio, idx, pr, in_stock, curl
        if best >= MATCH_THRESHOLD and bidx is not None:
            matched[our_norm] = (best, bidx, bprice, bstock, burl)
    def _base_of(n):
        return n.split(" - ")[0].strip() if " - " in n else n
    groups = {}
    for on, (sc, idx, pr, stock, url) in matched.items():
        groups.setdefault((_base_of(on), idx), []).append(on)
    drop = set()
    for (b, idx), members in groups.items():
        if len(members) > 1:
            members.sort(key=lambda on: (pack_count(on), on))
            drop.update(members[1:])
    return {on: (pr, (url or base), stock) for on, (sc, idx, pr, stock, url) in matched.items() if on not in drop}

_STORE_API_JS = r"""async (pg) => {
  try {
    const r = await fetch(location.origin + '/wp-json/wc/store/v1/products?per_page=100&page=' + pg, {headers:{'Accept':'application/json'}});
    if (!r.ok) return {err: r.status};
    const j = await r.json();
    if (!Array.isArray(j)) return {err: 'na'};
    return {items: j.map(p => ({n: p.name, price: (p.prices ? p.prices.price : null),
                                minor: (p.prices ? p.prices.currency_minor_unit : 2),
                                stock: p.is_in_stock, link: (p.permalink||''), slug: (p.slug || (p.permalink||'').replace(/\/+$/,'').split('/').pop())}))};
  } catch(e) { return {err: String(e)}; }
}"""

def collect_via_store_api(comp, product_norms):
    """WooCommerce Store API를 curl_cffi(크롬 TLS 흉내)로 직접 호출해 상품 전체를 JSON으로 수집.
    브라우저 불필요(창X, CAPTCHA X). meamoshop처럼 순수 requests는 막지만 크롬 지문엔 열리는 사이트용.
    실패시 None → 기존 카탈로그 방식으로 폴백."""
    base = comp["base"]
    catalog, seen = [], set()
    for pg in range(1, 80):
        txt = None
        for _try in range(4 if pg == 1 else 2):   # 1페이지(핵심)는 넉넉히 재시도(일시적 실패 대비)
            txt = cfetch(base + "/wp-json/wc/store/v1/products?per_page=100&page=%d" % pg, tries=2)
            if txt:
                break
            time.sleep(1.5)
        if not txt:
            break
        try:
            arr = json.loads(txt)
        except Exception:
            break
        if not isinstance(arr, list) or not arr:
            break
        for pjson in arr:
            prices = pjson.get("prices") or {}
            price = prices.get("price")
            if price in (None, ""):
                continue
            nm = html.unescape(pjson.get("name") or "")
            nn = normalize(nm)
            if not nn or nn in seen:
                continue
            try:
                pv = float(price) / (10 ** int(prices.get("currency_minor_unit", 2)))
            except Exception:
                continue
            if pv <= 0:
                continue
            seen.add(nn)
            link = pjson.get("permalink") or base
            instock = bool(pjson.get("is_in_stock", True))
            catalog.append((nn, unit_sig(nn), pv, instock, link))
            sl = (pjson.get("slug") or "").lower()
            if sl:
                _API_SLUG.setdefault(comp["key"], {})[sl] = (pv, instock)
        if len(arr) < 100:
            break
    if not catalog:
        return None
    result = _match_catalog(catalog, product_norms, base)
    SITE_STATUS[comp["key"]] = f"정상 (제품 {len(result)}개 매칭) [API-curl(빠름·브라우저X)·목록 {len(catalog)}개]"
    return result

def _api_by_slug(comp_key, slugs):
    """Store API를 slug로 curl_cffi 조회 → {slug:{'price','instock'}} (브라우저 없이)."""
    comp = COMP_BY_KEY.get(comp_key, {}); base = comp.get("base")
    out = {}
    if not base:
        return out
    for sslug in slugs:
        if not sslug:
            continue
        txt = cfetch(base + "/wp-json/wc/store/v1/products?slug=" + urllib.parse.quote(sslug))
        if not txt:
            continue
        try:
            arr = json.loads(txt)
        except Exception:
            continue
        if isinstance(arr, list) and arr:
            prices = (arr[0].get("prices") or {})
            price = prices.get("price")
            if price not in (None, ""):
                try:
                    v = float(price) / (10 ** int(prices.get("currency_minor_unit", 2)))
                    if v > 0:
                        out[sslug] = {"price": v, "instock": bool(arr[0].get("is_in_stock", True))}
                except Exception:
                    pass
    return out


def collect_via_catalog(comp, product_norms):
    catalog = []
    seen_names = set()
    delay = comp.get("cat_delay", 0.5)         # Cloudflare 있는 곳(meamoshop)만 크게
    cat_pages = comp.get("cat_pages", [])
    # --- 모드 결정: 먼저 requests로 찔러보고, 막히거나 상품이 안 나오면 그때만 브라우저 ---
    use_browser = False   # v9.45: 브라우저 미사용(requests+curl_cffi로 충분)
    def _fetch(u):
        if use_browser:
            return browser_fetch(u, wait_selector=comp.get("wait_selector"))
        return fetch(u)
    for ci, cat in enumerate(cat_pages):
        if ci > 0:
            time.sleep(delay if use_browser else SLEEP_BETWEEN)
        # 페이지네이션 따라가며 모든 페이지 수집(무한스크롤/더보기 사이트는 browser_fetch가 처리)
        page_url, visited, pages = cat, set(), 0
        while page_url and pages < MAX_CAT_PAGES:
            visited.add(page_url)
            html = _fetch(page_url)
            pages += 1
            if not html or is_blocked(html):
                break
            for nm, pr, in_stock, purl in parse_catalog(html):
                nn = normalize(nm)
                if nn and nn not in seen_names:
                    seen_names.add(nn)
                    full = urllib.parse.urljoin(comp["base"], purl) if purl else comp["base"]
                    catalog.append((nn, unit_sig(nn), pr, in_stock, full))
            page_url = _next_page_url(html, page_url, visited)
            if page_url:
                time.sleep(max(delay * 0.5, 0.5) if use_browser else SLEEP_BETWEEN)
    comp["_used_browser"] = use_browser   # 상태표시에 사용
    if use_browser:                       # 브라우저가 Cloudflare 통과한 쿼키를 requests세션에 복사(수동보정 병렬화용)
        try:
            _nc = _sync_browser_cookies()
            if _nc:
                print(f"    · {comp['key']}: 브라우저 쿼키 {_nc}개를 requests에 복사(수동보정 빠르게)", flush=True)
        except Exception:
            pass
    if not catalog:
        SITE_STATUS[comp["key"]] = "차단됨 (봇 차단 - 자동 접속 거부)"
        return {}
    result = _match_catalog(catalog, product_norms, comp["base"])
    SITE_STATUS[comp["key"]] = (f"정상 (제품 {len(result)}개 매칭) "
                                f"[{'브라우저' if comp.get('_used_browser') else 'requests(빠름)'}·목록 {len(catalog)}개]")
    return result

def load_products_from_csv():
    prods = []
    with open(PRODUCTS_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            prods.append({"name": row["our_product"].strip(),
                          "cost": row.get("our_cost", "").strip(),
                          "price": row.get("our_price", "").strip(),
                          "norm": normalize(row["our_product"])})
    return prods

def load_products_from_site():
    """WooCommerce Store API(JSON)로 우리 제품을 한 번에 수집. 동시요청 없음, 항상 최신, 신제품 자동 포함."""
    base = "https://filler-outlet.com"
    prods, seen = [], set()
    for pg in range(1, 80):
        raw = fetch(base + "/wp-json/wc/store/v1/products?per_page=100&page=" + str(pg))
        if not raw:
            break
        try:
            arr = json.loads(raw)
        except Exception:
            break
        if not isinstance(arr, list) or not arr:
            break
        for pr in arr:
            nm = html.unescape(pr.get("name") or "").strip()
            nn = normalize(nm)
            if not nn or nn in seen:
                continue
            prices = pr.get("prices") or {}
            try:
                minor = int(prices.get("currency_minor_unit", 2) or 2)
            except Exception:
                minor = 2
            raw = prices.get("price")
            if raw in (None, ""):
                rng = prices.get("price_range") or {}
                raw = rng.get("min_amount") if isinstance(rng, dict) else None
            try:
                price = float(raw) / (10 ** minor) if raw not in (None, "") else ""
            except Exception:
                price = ""
            seen.add(nn)
            prods.append({"name": nm, "cost": "", "price": price, "norm": nn})
        if len(arr) < 100:
            break
    print("  filler-outlet.com Store API로 " + str(len(prods)) + "개 수집")
    return prods

def load_extra_products():
    """'우리제품추가' 시트(제품명·판매가[·원가])를 읽어 우리 제품에 합칠 목록으로 반환.
    자동수집이 못 보는 로그인 가려진 제품(보톡스 등)용."""
    out = []
    if not OUR_PRODUCTS_SHEET_URL:
        return out
    text = fetch(OUR_PRODUCTS_SHEET_URL)
    if not text:
        print("  ! 우리제품추가 시트 로드 실패(무시)")
        return out
    def _f(x):
        try:
            return float(str(x).replace(",", "").replace("$", "").strip())
        except Exception:
            return ""
    try:
        for r in csv.DictReader(io.StringIO(text)):
            name = (r.get("제품명") or r.get("product") or "").strip()
            price = _f(r.get("판매가") or r.get("price") or "")
            cost = _f(r.get("원가") or r.get("cost") or "")
            if name:
                out.append({"name": name, "cost": cost, "price": price, "norm": normalize(name)})
    except Exception as e:
        print("  ! 우리제품추가 읽기 오류:", e)
    return out

def _merge_extra(prods):
    extra = load_extra_products()
    seen = {p["norm"] for p in prods}
    added = 0
    for e in extra:
        if e["norm"] and e["norm"] not in seen:
            prods.append(e); seen.add(e["norm"]); added += 1
    if added:
        print(f"  + 우리제품추가(시트)에서 {added}개 병합(보톡스 등)")
    return prods

def _split_elasty_options(prods):
    """Elasty G/F/D Plus(리도카인)는 filler-outlet 사이트에서 1실린지 가격 하나로 뭉쳐서 옴.
    → '- 1x1ml'(사이트가격)와 '- 2x1ml'(2배)로 쪼개, 경쟁사와 실린지 수량별로 정확히 매칭되게 한다.
    ('elasty g'처럼 이미 옵션/용량이 붙은 이름은 건드리지 않음)"""
    pat = re.compile(r"^\s*elasty\s+[gfd]\s+plus\s*$", re.I)
    out, seen = [], set()
    for p in prods:
        base = re.sub(r"\s+", " ", (p.get("name") or "")).strip()
        if pat.match(base) and isinstance(p.get("price"), (int, float)):
            for suffix, price in ((" - 1x1ml", p["price"]), (" - 2x1ml", round(p["price"] * 2, 2))):
                nm = base + suffix
                nn = normalize(nm)
                if nn in seen:
                    continue
                seen.add(nn)
                out.append({"name": nm, "cost": p.get("cost", ""), "price": price, "norm": nn})
        else:
            p = dict(p, name=_fix_dash(p.get("name") or ""))
            out.append(p)
            if p.get("norm"):
                seen.add(p["norm"])
    return out

def load_products():
    if PRODUCT_SOURCE == "site":
        try:
            print("우리 제품 목록: filler-outlet.com 자동 수집 중...")
            prods = load_products_from_site()
            if prods:
                try:
                    with open(PRODUCTS_CSV, "w", encoding="utf-8-sig", newline="") as f:
                        w = csv.writer(f); w.writerow(["our_product", "our_cost", "our_price"])
                        for p in prods:
                            w.writerow([p["name"], p["cost"], p["price"]])
                except Exception:
                    pass
                print(f"  -> 자동수집 {len(prods)}개")
                return _split_elasty_options(_merge_extra(prods))
            print("  ! 자동수집 실패 -> products.csv 사용")
        except Exception as e:
            print("  ! 자동수집 오류 -> products.csv 사용:", e)
    return _split_elasty_options(_merge_extra(load_products_from_csv()))

def collect_competitor(comp, product_norms):
    result = {}
    use_browser = bool(comp.get("browser")) and USE_BROWSER
    # WooCommerce Store API 우선(meamo 등): curl_cffi로 브라우저 없이 전체 JSON 수집
    if comp.get("store_api"):
        for _ in range(2):                     # 일시적 실패 대비 재시도
            try:
                res = collect_via_store_api(comp, product_norms)
                if res:
                    return res
            except Exception:
                pass
            time.sleep(2)
        # store_api 사이트는 브라우저로 폴백하지 않음(창+CAPTCHA 방지).
        #  이번 회차 카탈로그(자동매칭)만 비지만, 수동보정 meamo URL은 curl_cffi로 여전히 커버됨.
        SITE_STATUS[comp["key"]] = "정상 (store_api 일시실패-이번 회차 카탈로그 생략, 수동보정으로 커버)"
        return {}
    if comp.get("cat_pages"):
        return collect_via_catalog(comp, product_norms)
    fetcher = browser_fetch if use_browser else fetch
    urls = get_sitemap_urls(comp.get("sitemaps", []), fetcher=fetcher)
    if not urls and comp.get("categories"):
        urls = product_urls_from_pages(comp["categories"], comp["base"], fetcher=fetcher)
    url_slugs = [(u, normalize(slug_of(u)), unit_sig(normalize(slug_of(u)))) for u in urls]
    targets = {}
    for our_norm in product_norms:
        osig = unit_sig(our_norm)
        best, best_url = 0.0, None
        for u, sn, ssig in url_slugs:
            if not sn or not units_ok(osig, ssig) or not plus_ok(our_norm, sn) \
               or not pack_ok(our_norm, sn) or not _distinctive_ok(our_norm, sn):
                continue
            ratio = match_score(our_norm, sn)
            if ratio > best:
                best, best_url = ratio, u
        if best >= MATCH_THRESHOLD and best_url:
            targets.setdefault(best_url, our_norm)

    items = list(targets.items())
    blocked_n = 0
    if use_browser:
        # 브라우저는 순차 처리(스레드 불가)
        for url, our_norm in items:
            html = browser_fetch(url); time.sleep(0.2)
            if is_blocked(html):
                blocked_n += 1; continue
            price = parse_price_from_html(html) if html else None
            if price:
                result[our_norm] = (price, url, not is_out_of_stock(html))
    else:
        def worker(item):
            url, our_norm = item
            html = fetch(url); time.sleep(SLEEP_BETWEEN)
            if is_blocked(html):
                return our_norm, "__BLOCKED__"
            if not html:
                return our_norm, None
            price = parse_price_from_html(html)
            return our_norm, ((price, url, not is_out_of_stock(html)) if price else None)
        with cf.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            for our_norm, val in ex.map(worker, items):
                if val == "__BLOCKED__":
                    blocked_n += 1
                elif val:
                    result[our_norm] = val

    if not result and (blocked_n > 0 or not urls):
        probe = fetcher((comp.get("sitemaps") or [comp["base"]])[0]) or fetcher(comp["base"])
        if is_blocked(probe) or blocked_n > 0:
            SITE_STATUS[comp["key"]] = "차단됨 (봇 차단 - 자동 접속 거부)"
            return {}
    SITE_STATUS[comp["key"]] = f"정상 (제품 {len(result)}개 매칭)" + (" [브라우저]" if use_browser else "")
    return result

def load_previous_history():
    if not os.path.isdir(HISTORY_DIR):
        return None
    files = sorted(f for f in os.listdir(HISTORY_DIR) if f.endswith(".json"))
    if not files:
        return None
    with open(os.path.join(HISTORY_DIR, files[-1]), encoding="utf-8") as f:
        return json.load(f)

def save_history(catalogs, today):
    slim = {k: {n: v[0] for n, v in d.items()} for k, d in catalogs.items()}
    with open(os.path.join(HISTORY_DIR, f"{today}.json"), "w", encoding="utf-8") as f:
        json.dump(slim, f, ensure_ascii=False)

def _suspicious_keys(comp_prices):
    """다른 경쟁사 가격 대비 혼자 너무 싸거나(중앙값*0.4 미만) 너무 비싼(*3배 초과) 값 → 의심."""
    vals = [(k, v) for k, v in comp_prices.items() if isinstance(v, (int, float)) and v > 0]
    if len(vals) < 3:
        return set()
    nums = sorted(v for _, v in vals)
    n = len(nums)
    median = nums[n // 2] if n % 2 else (nums[n // 2 - 1] + nums[n // 2]) / 2
    if median <= 0:
        return set()
    return {k for k, v in vals if v < median * 0.4 or v > median * 3.0}

def _we_sell_match(bname, our_prods):
    """베스트셀러(경쟁사) 제품명이 우리 판매목록에 있는지 매칭. 있으면 우리 제품 dict, 없으면 None.
    가격비교와 동일 기준(용량·리도카인·실린지수량 구분)."""
    bn = normalize(bname or "")
    if not bn:
        return None
    bsig = unit_sig(bn)
    best, bestp = 0.0, None
    for op in our_prods:
        on = op.get("norm") or ""
        if not on or not units_ok(bsig, unit_sig(on)) or not plus_ok(bn, on) \
           or not pack_ok(bn, on) or not _distinctive_ok(bn, on):
            continue
        sc = match_score(bn, on)
        if sc > best:
            best, bestp = sc, op
    return bestp if best >= MATCH_THRESHOLD else None

def write_excel(rows, newly, today, bestsellers=None):
    wb = openpyxl.Workbook()
    fill = PatternFill("solid", fgColor="2F5496")
    hlpink = PatternFill("solid", fgColor="FFC7CE")   # 우리 등수 칸
    yel = PatternFill("solid", fgColor="FFF2CC")       # 의심 칸
    _unit_notes = load_unit_notes()
    def _unit_note(nm):
        lk = _loose_key(nm)
        return "단위주의(박스vs바이알)" if any(u and u in lk for u in _unit_notes) else ""
    def _num(x):
        return x if isinstance(x, (int, float)) else None
    def _memo_of(p):
        m = _unit_note(p["name"])
        uf = [k for k in COMP_KEYS if (k, p["norm"]) in URL_FAIL]
        if uf:
            m = (m + " / " if m else "") + "URL실패:" + ",".join(uf)
        return m

    # ===== 등수 계산 =====
    rank_rows, max_n = [], 0
    site_rank_cnt = {}
    for r in rows:
        p = r["p"]; susp = _suspicious_keys(r["comp"])
        entries = []   # (price, is_ours, comp_key, soldout, suspicious)
        op = _num(p["price"]) if str(p.get("price", "")).strip() != "" else None
        if op is not None:
            entries.append((op, True, "filler-outlet", False, False))
        for k in COMP_KEYS:
            v = _num(r["comp"].get(k))
            if v is not None:
                soldout = not r.get("stock", {}).get(k, True)
                entries.append((v, False, k, soldout, (k in susp)))
        entries.sort(key=lambda e: e[0])
        our_rank = next((i + 1 for i, e in enumerate(entries) if e[1]), None)
        for i, e in enumerate(entries):
            d = site_rank_cnt.setdefault(e[2], {})
            d[i + 1] = d.get(i + 1, 0) + 1
        max_n = max(max_n, len(entries))
        rank_rows.append((p, entries, our_rank, _memo_of(p)))
    max_n = min(max_n, 12) if max_n else 1
    K = min(max_n, 10)

    # ===== 시트1: 순위 요약 (사이트별 등수 분포) =====
    wsum = wb.create_sheet("순위 요약")
    wsum.append(["사이트명"] + [f"{i}등" for i in range(1, K + 1)])
    for c in wsum[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.alignment = Alignment(horizontal="center")
    for site in (COMP_KEYS[:3] + ["filler-outlet"] + COMP_KEYS[3:]):
        cnts = site_rank_cnt.get(site, {})
        label = "filleroutlet(우리)" if site == "filler-outlet" else site
        wsum.append([label] + [cnts.get(i, 0) for i in range(1, K + 1)])
        if site == "filler-outlet":
            for c in wsum[wsum.max_row]:
                c.fill = hlpink
    wsum.column_dimensions["A"].width = 22
    for i in range(1, K + 1):
        wsum.column_dimensions[get_column_letter(i + 1)].width = 8
    wsum.freeze_panes = "B2"

    # ===== 시트2: 순위표 (제품별) =====
    ws = wb.active; ws.title = "순위표"
    header = ["제품명", "판매가", "우리 등수", "메모"] + [f"{i}등" for i in range(1, max_n + 1)]
    NF = 4
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.alignment = Alignment(horizontal="center")
    for p, entries, our_rank, memo in rank_rows:
        line = [p["name"], p["price"], (our_rank if our_rank else ""), memo]
        for i in range(max_n):
            if i < len(entries):
                pr, ours, ck, soldout, suspicious = entries[i]
                if soldout or suspicious:
                    line.append(f"{pr}" + (" (품절)" if soldout else "") + (" ⚠" if suspicious else ""))
                else:
                    line.append(pr)
            else:
                line.append("")
        ws.append(line); rn = ws.max_row
        if our_rank and our_rank <= max_n:
            ws.cell(rn, NF + our_rank).fill = hlpink
        for i in range(min(max_n, len(entries))):
            if entries[i][4]:
                ws.cell(rn, NF + 1 + i).fill = yel
        if memo:
            ws.cell(rn, 4).font = Font(color="C55A11", bold=True)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["D"].width = 26
    for i in range(1, len(header)):
        col = get_column_letter(i + 1)
        if col not in ("A", "D"):
            ws.column_dimensions[col].width = 12
    ws.freeze_panes = "B2"

    # ===== 시트2: raw data (우리 쇼핑몰 filler-outlet 포함) =====
    ws5 = wb.create_sheet("raw data")
    header5 = ["제품명", "경쟁사", "링크", "가격", "메모"]   # 품절/의심/단위주의/URL실패를 메모 한 칸에 통합
    ws5.append(header5)
    for c in ws5[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.alignment = Alignment(horizontal="center")
    for r in rows:
        p = r["p"]; susp = _suspicious_keys(r["comp"])
        if str(p.get("price", "")).strip() != "":     # 우리 쇼핑몰 행(링크 없음)
            ws5.append([p["name"], "filler-outlet", "", p["price"], _unit_note(p["name"])])
        for k in COMP_KEYS:
            v = r["comp"].get(k)
            if v in (None, ""):
                continue
            url = r.get("url", {}).get(k, "")
            _parts = []
            if not r.get("stock", {}).get(k, True):
                _parts.append("품절")
            if k in susp:
                _parts.append("의심")
            _un = _unit_note(p["name"])
            if _un:
                _parts.append(_un)
            if (k, p["norm"]) in URL_FAIL:
                _parts.append("URL실패-확인필요")
            memo_c = " / ".join(_parts)
            ws5.append([p["name"], k, url, v, memo_c])
            if url and re.match(r"https?://[^/]+/.+", url):
                lc = ws5.cell(ws5.max_row, 3); lc.hyperlink = url
                lc.font = Font(color="0563C1", underline="single")
    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["B"].width = 16
    ws5.column_dimensions["C"].width = 50
    ws5.column_dimensions["D"].width = 10
    ws5.column_dimensions["E"].width = 32
    ws5.freeze_panes = "A2"

    # ===== 시트3: 사이트 상태 =====
    ws3 = wb.create_sheet("사이트 상태")
    ws3.append(["경쟁사", "상태 (자동 감지)"])
    for c in ws3[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
    warn = Font(color="C00000", bold=True)
    for k in COMP_KEYS:
        st = SITE_STATUS.get(k, "")
        ws3.append([k, st])
        if "차단" in st or "오류" in st:
            ws3.cell(ws3.max_row, 2).font = warn
    ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 50

    # ===== 시트4: 제외 목록 =====
    try:
        _ex = [(nm, cp) for (nm, cp, u, pr_) in load_overrides() if not (u or "").strip() and not (pr_ or "").strip()]
    except Exception:
        _ex = []
    ws6 = wb.create_sheet("제외 목록")
    ws6.append(["제품명", "제외한 경쟁사", "메모(왜 뺐나 / 나중에 확인)"])
    for c in ws6[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.alignment = Alignment(horizontal="center")
    if _ex:
        for nm, cp in _ex:
            ws6.append([nm, cp, "경쟁사가 우리 제품 미취급/가짜매칭이라 제외 — 나중에 실제 취급 시 수동보정에 URL 추가"])
    else:
        ws6.append(["(제외 처리된 항목 없음)", "", ""])
    ws6.column_dimensions["A"].width = 30
    ws6.column_dimensions["B"].width = 18
    ws6.column_dimensions["C"].width = 60

    # ===== 베스트셀러 시트 (세로 정리: 사이트별로 쭉) =====
    if bestsellers:
        wsb = wb.create_sheet("베스트셀러")
        bkeys = [k for k in COMP_KEYS if bestsellers.get(k)]
        nodata = [k for k in COMP_KEYS if not bestsellers.get(k)]
        note = "※ 베스트셀러 자료 없음(사이트가 판매순을 공개 안 함): " + (", ".join(nodata) if nodata else "없음")
        wsb.append([note])
        wsb.cell(1, 1).font = Font(bold=True, color="C00000")
        wsb.append([])
        wsb.append(["사이트명", "등수", "제품명", "판매가격", "우리판매", "매칭된 우리제품"])
        _our_prods = [r["p"] for r in rows]
        _hr = wsb.max_row
        for c in wsb[_hr]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
            c.alignment = Alignment(horizontal="center")
        _alt = PatternFill("solid", fgColor="EEF2FA")
        _bandi = 0
        for k in bkeys:
            lst = bestsellers.get(k, [])
            if not lst:
                continue
            _start = wsb.max_row + 1
            for i, (nm_b, pr_b) in enumerate(lst):
                _mp = _we_sell_match(nm_b, _our_prods)
                wsb.append([k, i + 1, nm_b, pr_b, ("O 판매중" if _mp else "X 미판매"), (_mp["name"] if _mp else "-")])
                if not _mp:
                    wsb.cell(wsb.max_row, 5).font = Font(bold=True, color="C00000")
            _end = wsb.max_row
            wsb.merge_cells(start_row=_start, start_column=1, end_row=_end, end_column=1)
            _sc = wsb.cell(_start, 1)
            _sc.alignment = Alignment(horizontal="center", vertical="center")
            _sc.font = Font(bold=True)
            if _bandi % 2 == 1:
                for _rr in range(_start, _end + 1):
                    wsb.cell(_rr, 1).fill = _alt
            _bandi += 1
        wsb.column_dimensions["A"].width = 20
        wsb.column_dimensions["B"].width = 6
        wsb.column_dimensions["C"].width = 42
        wsb.column_dimensions["D"].width = 12
        wsb.column_dimensions["E"].width = 11
        wsb.column_dimensions["F"].width = 34
        wsb.freeze_panes = "A" + str(_hr + 1)

    # 모든 시트 셀에 얇은 회색 테두리(구분 잘 되게)
    _thin = Side(style="thin", color="000000")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _ws in wb.worksheets:
        if _ws.max_row < 1:
            continue
        for _row in _ws.iter_rows(min_row=1, max_row=_ws.max_row, max_col=_ws.max_column):
            for _cell in _row:
                _cell.border = _border

    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out = os.path.join(OUT_DIR, f"경쟁사_가격비교_{stamp}.xlsx")
    try:
        wb.save(out)
        print(f"\n엑셀 저장: {out}")
    except PermissionError:
        alt = os.path.join(OUT_DIR, f"경쟁사_가격비교_{stamp}_2.xlsx")
        wb.save(alt)
        print(f"\n엑셀 저장: {alt}")

COMP_BY_KEY = {c["key"]: c for c in COMPETITORS}
_API_SLUG = {}   # comp_key -> {slug: (price, in_stock)}  (meamo API 수집 시 채움)

def _slug_of(url):
    return (url or "").split("?")[0].rstrip("/").split("/")[-1].lower()
OVERRIDES_CSV = os.path.join(HERE, "수동보정.csv")
UNIT_NOTE_FILE = os.path.join(HERE, "단위주의.txt")

# ── 구글 시트 중앙관리(선택) ─────────────────────────────────────────
# 아래 URL을 채우면 수동보정/단위주의를 '구글 시트'에서 읽는다(로컬 파일보다 우선).
# 시트를 고치면 그 다음 실행부터 모든 사람에게 자동 반영됨. 비우면 로컬 파일 사용.
# URL 형식(시트를 '링크가 있는 모든 사용자: 뷰어'로 공유한 뒤):
#   https://docs.google.com/spreadsheets/d/<시트ID>/gviz/tq?tqx=out:csv&sheet=<탭이름>
OVERRIDES_SHEET_URL = "https://docs.google.com/spreadsheets/d/1FO6QV6Vfxw8sndcaCrZKMAnsyuk5ogBl-IDxKnDSFgA/gviz/tq?tqx=out:csv&gid=0"
UNIT_NOTE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1FO6QV6Vfxw8sndcaCrZKMAnsyuk5ogBl-IDxKnDSFgA/gviz/tq?tqx=out:csv&gid=331662492"
# '우리제품추가' 탭: 자동수집이 못 보는 우리 제품(로그인 가려진 보톡스 등)을 제품명·판매가로 추가.
OUR_PRODUCTS_SHEET_URL = "https://docs.google.com/spreadsheets/d/1FO6QV6Vfxw8sndcaCrZKMAnsyuk5ogBl-IDxKnDSFgA/gviz/tq?tqx=out:csv&gid=889442535"

def _read_source(sheet_url, local_path, label):
    """구글시트 URL이 있으면 거기서, 없거나 실패하면 로컬 파일에서 텍스트를 읽어 반환."""
    if sheet_url:
        txt = fetch(sheet_url)
        if txt and ("," in txt or "\n" in txt):
            print(f"  {label}: 구글시트에서 로드", flush=True)
            return txt
        print(f"  ! {label}: 구글시트 로드 실패 → 로컬 파일 사용", flush=True)
    if os.path.isfile(local_path):
        try:
            return open(local_path, encoding="utf-8-sig").read()
        except Exception:
            return None
    return None

def load_unit_notes():
    """단위주의.txt: 한 줄에 제품명 하나. 우리(박스)와 경쟁사(바이알)처럼 단위가 달라
    가격을 그대로 비교하면 오해가 생기는 제품 목록. 엑셀에 '단위주의' 메모로 표시된다.
    빈 줄·# 로 시작하는 줄은 무시."""
    notes = []
    text = _read_source(UNIT_NOTE_SHEET_URL, UNIT_NOTE_FILE, "단위주의")
    if not text:
        return notes
    for row in csv.reader(io.StringIO(text)):   # 시트(CSV)·txt 둘 다: 각 행 첫 칸을 제품명으로
        if not row:
            continue
        t = (row[0] or "").strip()
        if not t or t.startswith("#") or t.lower() in ("제품명", "product", "name"):
            continue
        notes.append(_loose_key(t))
    return notes

def _loose_key(s):
    return re.sub(r"[\s\-–—]", "", normalize(s))

def load_overrides():
    """수동보정: 제품명, 경쟁사, URL, 가격 (URL 우선/가격 대체/둘다 비면 제외).
    OVERRIDES_SHEET_URL 있으면 구글시트에서, 없으면 로컬 수동보정.csv에서 읽음."""
    out = []
    text = _read_source(OVERRIDES_SHEET_URL, OVERRIDES_CSV, "수동보정")
    if not text:
        return out
    try:
        for r in csv.DictReader(io.StringIO(text)):
            name = (r.get("제품명") or r.get("product") or "").strip()
            comp = (r.get("경쟁사") or r.get("competitor") or "").strip()
            url  = (r.get("URL") or r.get("url") or "").strip()
            price = (r.get("가격") or r.get("price") or "").strip()
            if name and comp:
                out.append((name, comp, url, price))
    except Exception as e:
        print("  ! 보정 읽기 오류:", e)
    return out

def _override_fetch(comp_key, url):
    comp = COMP_BY_KEY.get(comp_key, {})
    html = fetch(url)
    if (not html or is_blocked(html)) and comp.get("browser") and USE_BROWSER:
        html = browser_fetch(url)
    return html

_ESTA_SEARCH = "https://estaderma.com/shop/?search="
_esta_search_cache = {}
def _estaderma_price_by_search(onorm):
    """estaderma 상세URL이 로그인(401)으로 막힐 때, 공개 검색페이지에서 가격을 찾아 매칭."""
    parts = onorm.split()
    term = parts[0] if parts else onorm
    if len(term) < 3:
        return None
    if term not in _esta_search_cache:
        q = _ESTA_SEARCH + urllib.parse.quote(term)
        try:
            html = cfetch(q) or fetch(q)
        except Exception:
            html = None
        cat = []
        if html and not is_blocked(html):
            try:
                cat = parse_catalog(html)
            except Exception:
                cat = []
        _esta_search_cache[term] = cat
    best, bprice, bstock = 0.0, None, True
    osig = unit_sig(onorm)
    for t in _esta_search_cache[term]:
        nm = normalize(str(t[0]))
        if not nm or not units_ok(osig, unit_sig(nm)) or not plus_ok(onorm, nm) or not pack_ok(onorm, nm):
            continue
        sc = match_score(onorm, nm)
        if sc > best:
            best, bprice, bstock = sc, t[1], (t[2] if len(t) > 2 else True)
    if best >= MATCH_THRESHOLD and bprice:
        return (bprice, bstock)
    return None

def apply_overrides(catalogs, products):
    """수동보정.csv(제품명·경쟁사·URL·가격) 적용. URL 있으면 그 URL의 실시간 가격으로 덮어씀.
    meamo는 이미 API로 받아온 슬러그맵에서 조회(무접속), 나머지는 병렬 requests."""
    ov = load_overrides()
    if not ov:
        return
    by_loose = {}
    for p in products:
        by_loose.setdefault(_loose_key(p["name"]), p["norm"])
    items = []
    for name, comp, url, price in ov:
        if comp not in COMP_KEYS:
            continue
        onorm = by_loose.get(_loose_key(name)) or normalize(name)
        items.append((onorm, comp, (url or "").strip(), (price or "").strip()))
    for onorm, comp, url, price in items:               # '제외'행(URL·가격 둘다 없음)만 자동값 제거
        if not url and not price:                       # URL 있는데 실패하면 아래서 자동매칭값 유지(폴백)
            catalogs.setdefault(comp, {}).pop(onorm, None)
    print(f"  수동보정: {len(items)}건 처리 중(URL 시세 갱신)...", flush=True)
    _ov_t0 = time.time()
    applied = 0
    meamo_slugs = _API_SLUG.get("meamoshop", {})
    fetch_tasks, meamo_browser = [], []
    for onorm, comp, url, price in items:
        if url:
            if comp == "meamoshop" and meamo_slugs:
                hit = meamo_slugs.get(_slug_of(url))
                if hit:
                    catalogs["meamoshop"][onorm] = (hit[0], url, hit[1]); applied += 1; continue
                meamo_browser.append((onorm, url)); continue
            fetch_tasks.append((onorm, comp, url))
        elif price:
            try:
                pv = float(price.replace(",", "").replace("$", ""))
                if pv > 0:
                    catalogs[comp][onorm] = (pv, "(수동가격)", True); applied += 1
            except Exception:
                pass
    _fail = {"blocked": 0, "parse": 0, "neterr": 0}   # 실패 원인 집계(진단용)
    def _work(t):
        onorm, comp, url = t
        reason = "neterr"
        for _ in range(2):                       # 일시적 실패 대비 재시도
            try:
                html = fetch(url)
                if not html:
                    reason = "neterr"; continue
                if is_blocked(html):
                    reason = "blocked"; continue   # Cloudflare 등 차단 → 브라우저 폴백 필요
                pr = parse_price_from_html(html)
                if pr:
                    return (onorm, comp, url, pr, not is_out_of_stock(html), None)
                reason = "parse"                   # 페이지는 받았는데 가격 못 읽음
            except Exception:
                reason = "neterr"
        return (onorm, comp, url, None, None, reason)
    browser_retry = []                           # requests 실패한 브라우저사이트(acecosm/mjs) → 브라우저 재시도
    _retry2 = []                                 # 1차 실패 → 2차 재시도 대상
    if fetch_tasks:
        with cf.ThreadPoolExecutor(max_workers=8) as ex:
            for onorm, comp, url, pr, instock, reason in ex.map(_work, fetch_tasks):
                if pr:
                    catalogs[comp][onorm] = (pr, url, instock); applied += 1
                else:
                    if reason in _fail:
                        _fail[reason] += 1
                    _retry2.append((onorm, comp, url))
                    if COMP_BY_KEY.get(comp, {}).get("browser") and USE_BROWSER:
                        browser_retry.append((onorm, comp, url))
    # --- 2차 재시도: 1차 실패(대개 부하 때 네트워크 타임아웃)만 동시 3개로 여유있게 다시 ---
    _retry2_ok = 0
    if _retry2:
        print(f"  · 1차 실패 {len(_retry2)}건 2차 재시도(여유있게)...", flush=True)
        def _work2(t):
            onorm, comp, url = t
            for _k in range(3):
                try:
                    html = fetch(url)
                    if html and not is_blocked(html):
                        pr = parse_price_from_html(html)
                        if pr:
                            return (onorm, comp, url, pr, not is_out_of_stock(html))
                except Exception:
                    pass
                time.sleep(2.0)
            return (onorm, comp, url, None, None)
        with cf.ThreadPoolExecutor(max_workers=3) as ex:
            for onorm, comp, url, pr, instock in ex.map(_work2, _retry2):
                if pr:
                    catalogs[comp][onorm] = (pr, url, instock); applied += 1; _retry2_ok += 1
        if _retry2_ok:
            print(f"    ✓ 2차 재시도로 {_retry2_ok}건 추가 확보", flush=True)
    # --- estaderma: 상세URL 401(로그인) 실패분을 공개 검색페이지로 폴백 ---
    _esta_ok = 0
    _esta_left = [(o, c, u) for (o, c, u) in _retry2 if c == "estaderma" and not catalogs.get(c, {}).get(o)]
    if _esta_left:
        print("  · estaderma " + str(len(_esta_left)) + "건 검색으로 폴백...", flush=True)
        for onorm, comp, url in _esta_left:
            hit = _estaderma_price_by_search(onorm)
            if hit:
                catalogs.setdefault(comp, {})[onorm] = (hit[0], url, hit[1]); applied += 1; _esta_ok += 1
        if _esta_ok:
            print("    ✓ estaderma 검색으로 " + str(_esta_ok) + "건 확보", flush=True)
    # --- 핵심: 브라우저 안에서 병렬 fetch (requests가 못뚫은 mjs 등) ---
    _batch_applied = 0
    if browser_retry and USE_BROWSER:
        by_comp = {}
        for onorm, comp, url in browser_retry:
            if COMP_BY_KEY.get(comp, {}).get("browser"):
                by_comp.setdefault(comp, []).append((onorm, url))
        still = []
        for comp, tlist in by_comp.items():
            print(f"  · {comp}: 브라우저 병렬 fetch {len(tlist)}건...", flush=True)
            res = {}
            try:
                res = _browser_batch_prices(comp, tlist)
            except Exception as _e:
                print(f"    (병렬 fetch 오류: {_e})", flush=True)
            for onorm, url in tlist:
                v = res.get(url)
                if v and v.get("price"):
                    catalogs[comp][onorm] = (v["price"], url, bool(v.get("instock", True)))
                    _batch_applied += 1; applied += 1
                else:
                    still.append((onorm, comp, url))
        browser_retry = still                       # 병렬로도 못 얻은 것만 순차폴백으로
        if _batch_applied:
            print(f"    ✓ 브라우저 병렬로 {_batch_applied}건 확보(순차폴백 {len(browser_retry)}건 남음)", flush=True)
    # 남은 것만 순차 재시도. 상한을 둔다(개수+시간).
    BROWSER_RETRY_CAP = 20            # 최대 재시도 개수
    BROWSER_RETRY_BUDGET = 180        # 최대 소요(초)
    _br_requested = len(browser_retry)
    _br_skipped = 0
    if len(browser_retry) > BROWSER_RETRY_CAP:
        _br_skipped = len(browser_retry) - BROWSER_RETRY_CAP
        browser_retry = browser_retry[:BROWSER_RETRY_CAP]
    _br_t0 = time.time()
    for onorm, comp, url in browser_retry:       # 순차 브라우저 재시도(스레드 불가)
        if time.time() - _br_t0 > BROWSER_RETRY_BUDGET:   # 시간초과 → 남은 건 스킵(전체 멈춤 방지)
            _br_skipped += (len(browser_retry) - browser_retry.index((onorm, comp, url)))
            break
        try:
            html = browser_fetch(url, tries=1)
            if html and not is_blocked(html):
                pr = parse_price_from_html(html)
                if pr:
                    catalogs[comp][onorm] = (pr, url, not is_out_of_stock(html)); applied += 1
        except Exception:
            pass
    # meamo 슬러그맵에 없던 것: 먼저 Store API를 slug로 직접 조회(빠름, 버림 없음)
    _meamo_api = 0
    if meamo_browser:
        _mslugs = [_slug_of(u) for (_on, u) in meamo_browser]
        try:
            _mres = _api_by_slug("meamoshop", _mslugs)
        except Exception:
            _mres = {}
        _still = []
        for onorm, url in meamo_browser:
            v = _mres.get(_slug_of(url))
            if v and v.get("price"):
                catalogs["meamoshop"][onorm] = (v["price"], url, bool(v.get("instock", True)))
                applied += 1; _meamo_api += 1
            else:
                _still.append((onorm, url))
        if _meamo_api:
            print(f"    ✓ meamo: Store API(slug조회)로 {_meamo_api}건 확보(개별 브라우저 {len(_still)}건만 남음)", flush=True)
        meamo_browser = _still
    # 그래도 못 얻은 극소수만 개별 브라우저(최후수단). 개수+시간 상한.
    MEAMO_BROWSER_CAP = 15
    MEAMO_BROWSER_BUDGET = 120
    if not USE_BROWSER:
        meamo_browser = []          # 브라우저 미사용 → 개별 브라우저 폴백 안함(잔여는 URL실패로 표시)
    _mb_total = len(meamo_browser)
    _mb_skipped = max(0, _mb_total - MEAMO_BROWSER_CAP)
    _mb_t0 = time.time()
    for onorm, url in meamo_browser[:MEAMO_BROWSER_CAP]:
        if time.time() - _mb_t0 > MEAMO_BROWSER_BUDGET:   # 시간초과 → 남은 건 스킵(25분 사태 방지)
            _mb_skipped += 1
            continue
        try:
            html = browser_fetch(url)
            if html and not is_blocked(html):
                pr = parse_price_from_html(html)
                if pr:
                    catalogs["meamoshop"][onorm] = (pr, url, not is_out_of_stock(html)); applied += 1
        except Exception:
            pass
    # 수동보정 URL이 있었는데 최종값이 그 URL에서 온 게 아니면(=조회 실패→자동폴백/빈칸) 표시용으로 기록
    URL_FAIL.clear()
    for onorm, comp, url, price in items:
        if url and comp in COMP_KEYS:
            cur = catalogs.get(comp, {}).get(onorm)
            if (not cur) or (len(cur) > 1 and cur[1] != url):
                URL_FAIL[(comp, onorm)] = True
    _ov_dt = time.time() - _ov_t0
    _br = _br_requested
    print(f"  ✓ 수동보정 {applied}/{len(items)}건 적용됨 (소요 {_ov_dt:.0f}초)")
    if _br or any(_fail.values()) or _batch_applied:
        _skip_txt = f", 시간/개수상한으로 스킵 {_br_skipped}건" if _br_skipped else ""
        _batch_txt = f", 브라우저 병렬확보 {_batch_applied}건" if _batch_applied else ""
        _meamo_api_txt = f", meamo슬러그API {_meamo_api}건" if _meamo_api else ""
        _meamo_txt = _meamo_api_txt + (f", meamo개별 {min(_mb_total, MEAMO_BROWSER_CAP)}건(스킵 {_mb_skipped})" if _mb_total else "")
        _r2txt = f", 2차재시도확보 {_retry2_ok}건" if _retry2_ok else ""
        print(f"    └ 진단: requests실패 {sum(_fail.values())}건"
              f"(차단 {_fail['blocked']}·파싱실패 {_fail['parse']}·네트워크 {_fail['neterr']}){_r2txt}"
              f"{_batch_txt}, 브라우저 순차재시도 {_br}건{_skip_txt}{_meamo_txt}")

BESTSELLER = {
    "meamoshop":          {"api": True},
    "koreafillerexperts": {"api": True},
    "derma-solution":     {"api": True},
    "estaderma":          {"url": "https://estaderma.com/shop/?orderby=popularity"},
    "fillerhouse":        {"url": "https://fillerhouse.com/shop/?price=&search=&sort=tmp_popularity"},
    # acecosm, mjsmedicals: 판매순(베스트셀러) 데이터를 공개 안 함 -> 시트에 "자료 없음" 표시
}

def _bs_noise(name):
    """베스트셀러 목록의 홍보/비제품 항목 걸러내기."""
    n = (name or "").strip().lower()
    if len(n) < 3:
        return True
    if re.fullmatch(r"[\d$.,\s%~\-]+", n):   # "100$" 처럼 숫자/기호만
        return True
    if n in ("hot", "new", "sale", "best", "hot deal", "best seller", "bestseller",
             "sale!", "hot!", "핫", "신상", "인기", "특가", "이벤트", "세일"):
        return True
    for w in ("hot new", "new arrival", "gift card", "e-gift", "coupon",
              "sample only", "clearance sale", "쿠폰", "기프트", "샘플"):
        if w in n:
            return True
    return False

def _collect_bestsellers_one(comp, top_n):
    key = comp["key"]; base = comp["base"]
    cfg = BESTSELLER.get(key)
    if not cfg:
        return key, []
    items, seen = [], set()
    def _add(nm, pr):
        nm = html.unescape(nm or "").strip()
        k2 = normalize(nm)
        if nm and k2 and k2 not in seen and not _bs_noise(nm):
            seen.add(k2); items.append((nm, pr))
    try:
        if cfg.get("api"):
            txt = cfetch(base + "/wp-json/wc/store/v1/products?orderby=popularity&per_page=" + str(top_n + 8))
            if txt:
                arr = json.loads(txt)
                if isinstance(arr, list):
                    for pp in arr:
                        prices = pp.get("prices") or {}
                        try:
                            minor = int(prices.get("currency_minor_unit", 2) or 2)
                            raw = prices.get("price")
                            price = float(raw) / (10 ** minor) if raw not in (None, "") else ""
                        except Exception:
                            price = ""
                        _add(pp.get("name"), price)
        elif cfg.get("url") or cfg.get("urls"):
            cands = cfg.get("urls") or [cfg.get("url")]
            for url in cands:
                if not url:
                    continue
                page = cfetch(url) or fetch(url)
                if (not page or is_blocked(page)) and cfg.get("browser"):
                    try:
                        page = browser_fetch(url, wait_selector=comp.get("wait_selector"))
                    except Exception:
                        page = None
                if page and not is_blocked(page):
                    for t in parse_catalog(page, include_sets=True):
                        _add(str(t[0]), t[1] if len(t) > 1 else "")
                if items:      # 이 주소에서 제품이 나왔으면 그만
                    break
    except Exception as e:
        print("  ! 베스트셀러(" + key + ") 오류:", str(e)[:60])
    return key, items[:top_n]

def collect_bestsellers(top_n=20):
    """경쟁사별 인기/베스트셀러 상위 N개. {key: [(제품명, 가격), ...]}.
    v9.67: 7개 사이트를 (PARALLEL_SITES면) 동시 수집 → 이 단계 대기시간 대폭 단축."""
    result = {}
    if PARALLEL_SITES:
        with cf.ThreadPoolExecutor(max_workers=len(COMPETITORS)) as ex:
            for key, items in ex.map(lambda c: _collect_bestsellers_one(c, top_n), COMPETITORS):
                result[key] = items
                print("  - 베스트셀러 " + key + ": " + str(len(items)) + "개", flush=True)
    else:
        for comp in COMPETITORS:
            key, items = _collect_bestsellers_one(comp, top_n)
            result[key] = items
            print("  - 베스트셀러 " + key + ": " + str(len(items)) + "개", flush=True)
    return result

def run():
    _run_t0 = time.time()
    print(f"=== 프로그램 버전: {VERSION} ===", flush=True)
    print(f"시작 시각: {datetime.datetime.now().strftime('%H:%M:%S')}", flush=True)
    os.makedirs(HISTORY_DIR, exist_ok=True)
    today = datetime.date.today().isoformat()
    products = load_products()
    product_norms = [p["norm"] for p in products]
    print(f"\n제품 {len(products)}개 준비. 경쟁사 {len(COMPETITORS)}곳 조사 시작...\n")
    catalogs = {}
    def _collect_one(comp):
        try:
            return comp["key"], (collect_competitor(comp, product_norms) or {}), None
        except Exception as e:
            return comp["key"], {}, str(e)
    if PARALLEL_SITES:
        print(f"  경쟁사 {len(COMPETITORS)}곳 동시 조사 중... (사이트별 부담은 그대로)", flush=True)
        with cf.ThreadPoolExecutor(max_workers=len(COMPETITORS)) as ex:
            for key, res, err in ex.map(_collect_one, COMPETITORS):
                catalogs[key] = res
                if err:
                    SITE_STATUS[key] = f"오류: {err}"
                print(f"  - {key}: {SITE_STATUS.get(key, '완료')}", flush=True)
    else:
        for comp in COMPETITORS:
            key, res, err = _collect_one(comp)
            catalogs[key] = res
            if err:
                SITE_STATUS[key] = f"오류: {err}"
            print(f"  - {key}: {SITE_STATUS.get(key, '완료')}", flush=True)
    try:
        apply_overrides(catalogs, products)   # 수동보정.csv 적용(브라우저 닫기 전)
    except Exception as e:
        print("  ! 보정 적용 오류:", e)
    try:
        print("\n베스트셀러 수집 중...")
        bestsellers = collect_bestsellers(20)
    except Exception as e:
        print("  ! 베스트셀러 수집 오류:", e)
        bestsellers = {}
    close_browser()
    print("\n===== 사이트별 상태 요약 =====")
    for k in COMP_KEYS:
        print(f"  {k:20} {SITE_STATUS.get(k, '?')}")
    print("=============================\n")
    prev = load_previous_history()
    rows, newly = [], []
    for p in products:
        n = p["norm"]; comp_prices = {}; comp_stock = {}; comp_url = {}
        for k in COMP_KEYS:
            v = catalogs.get(k, {}).get(n)
            if v:
                comp_prices[k] = v[0]
                comp_stock[k] = (v[2] if len(v) > 2 else True)
                if len(v) > 1 and v[1]:
                    comp_url[k] = v[1]
                if prev is not None and not prev.get(k, {}).get(n):
                    newly.append((p["name"], k))
        nums = [x for x in comp_prices.values() if isinstance(x, (int, float))]
        lowest = min(nums) if nums else None
        cheapest = ""
        if lowest is not None and str(p["price"]).strip():
            try:
                cheapest = "O" if float(p["price"]) <= lowest else "X"
            except ValueError:
                cheapest = ""
        rows.append({"p": p, "comp": comp_prices, "stock": comp_stock, "url": comp_url, "lowest": lowest, "cheapest": cheapest})
    write_excel(rows, newly, today, bestsellers)
    save_history(catalogs, today)
    _total = time.time() - _run_t0
    _m, _s = divmod(int(_total), 60)
    print(f"완료! (총 소요 {_m}분 {_s}초, 종료 {datetime.datetime.now().strftime('%H:%M:%S')})")

if __name__ == "__main__":
    run()
